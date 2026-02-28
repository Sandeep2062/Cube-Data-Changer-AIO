"""
Cube Data Processor Module
Reads generated data and populates an office template workbook.

Based on: https://github.com/Sandeep2062/Cube-Data-Processor
"""

import os
import shutil
import openpyxl

from generator import generate_rows, grade_display_name, MORTAR_TYPES, ALL_TYPES


# ── Helpers ─────────────────────────────────────────────────────────────────

def _normalise_grade_name(raw):
    """Normalise a grade string for matching (strip spaces, uppercase)."""
    return raw.replace(" ", "").upper()


def _extract_grade_from_filename(filepath):
    """Extract grade name from a grade-file filename (legacy support)."""
    name = os.path.basename(filepath).split(".")[0].upper()
    if "MORTAR" in name and "_" in name:
        parts = name.split("_")
        if len(parts) >= 3:
            return f"{parts[-2]}:{parts[-1]}"
    return name.replace("_", "").replace("-", "").strip()


def _load_workbook(filepath):
    """Open a workbook with safe defaults."""
    try:
        return openpyxl.load_workbook(filepath, keep_vba=False, data_only=False, keep_links=False)
    except Exception:
        return openpyxl.load_workbook(filepath)


def _find_sheets_for_grade(office_wb, grade_name, log):
    """Return list of sheet names whose B12 matches *grade_name*."""
    target = _normalise_grade_name(grade_name)
    matches = []
    for sheet_name in office_wb.sheetnames:
        ws = office_wb[sheet_name]
        b12 = ws["B12"].value
        if b12 and _normalise_grade_name(str(b12)) == target:
            matches.append(sheet_name)
    return matches


def _grade_from_template_cell(value):
    """Resolve grade/type from template cell B12. Returns None if unsupported."""
    if value is None:
        return None

    raw = str(value).strip().upper()
    if not raw:
        return None

    normalized = raw.replace(" ", "").replace("_", "").replace("-", "")

    # Concrete grades (M10-M45)
    concrete = {g for g in ALL_TYPES if g.startswith("M")}
    if normalized in concrete:
        return normalized

    # Mortar variants commonly seen in templates
    mortar_14_tokens = {"1:4", "1/4", "14", "MORTAR1:4", "MORTAR1/4", "MORTAR14"}
    mortar_16_tokens = {"1:6", "1/6", "16", "MORTAR1:6", "MORTAR1/6", "MORTAR16"}

    if normalized in mortar_14_tokens:
        return "1:4"
    if normalized in mortar_16_tokens:
        return "1:6"

    # Additional fuzzy support
    if "MORTAR" in normalized and ("1:4" in raw or "1/4" in raw or normalized.endswith("14")):
        return "1:4"
    if "MORTAR" in normalized and ("1:6" in raw or "1/6" in raw or normalized.endswith("16")):
        return "1:6"

    return None


# ── Calendar logic ──────────────────────────────────────────────────────────

def load_calendar_data(calendar_file, log):
    """Load calendar Excel → dict[casting_date_str] → {7_days, 28_days}."""
    if not calendar_file or not os.path.exists(calendar_file):
        log("⚠ No calendar file selected")
        return None

    try:
        wb = _load_workbook(calendar_file)
        ws = wb.active
        cal = {}
        row = 2
        while True:
            casting = ws.cell(row=row, column=1).value
            if not casting:
                break
            d7 = ws.cell(row=row, column=2).value
            d28 = ws.cell(row=row, column=3).value
            key = str(casting).strip()
            cal[key] = {
                "7_days": str(d7).strip() if d7 else "",
                "28_days": str(d28).strip() if d28 else "",
            }
            row += 1
        wb.close()
        log(f"✓ Calendar loaded: {len(cal)} dates")
        return cal
    except Exception as e:
        log(f"✖ Calendar error: {e}")
        return None


# ── Date processing ─────────────────────────────────────────────────────────

def apply_dates(office_wb, calendar_data, log):
    """Write 7-day/28-day dates into every sheet based on C17 casting date."""
    updated = 0
    for sheet_name in office_wb.sheetnames:
        ws = office_wb[sheet_name]
        casting = ws["C17"].value
        if not casting:
            continue
        key = str(casting).strip()
        if key in calendar_data:
            d7 = calendar_data[key]["7_days"]
            d28 = calendar_data[key]["28_days"]
            if d7:
                ws["C18"] = d7
            if d28:
                ws["F18"] = d28
            updated += 1
            log(f"  ✓ {sheet_name}: {key} → 7d:{d7}, 28d:{d28}")
        else:
            log(f"  ⚠ Date not in calendar: {key} ({sheet_name})")
    return updated


# ── Grade processing (in-memory generation) ─────────────────────────────────

def apply_generated_grades(office_wb, selected_grades, num_rows, log, progress_cb=None):
    """
    For each selected grade, generate data in-memory and write directly
    into matching sheets of the office workbook.

    Parameters
    ----------
    office_wb : openpyxl.Workbook
    selected_grades : list[str]      e.g. ["M20", "M25", "1:4"]
    num_rows : int                   rows to generate per grade (should >= sheets)
    log : callable
    progress_cb : callable(float)    optional 0-1 progress callback

    Returns total number of sheets populated.
    """
    total = 0
    grade_count = len(selected_grades)

    for gi, grade in enumerate(selected_grades):
        display = grade_display_name(grade)
        sheets = _find_sheets_for_grade(office_wb, grade, log)
        log(f"\n  Grade: {display}  →  {len(sheets)} matching sheets")

        if not sheets:
            log(f"  ⚠ No sheets with B12 = '{grade}'")
            continue

        rows_needed = len(sheets)
        gen = generate_rows(grade, rows_needed)

        for si, (weights, s7d, s28d) in enumerate(gen):
            if si >= len(sheets):
                break
            ws = office_wb[sheets[si]]

            # Weights → row 25, columns C-H (3-8)
            for i, v in enumerate(weights):
                ws.cell(row=25, column=3 + i, value=v)

            # 7-day strengths → row 27, columns C-E (3-5)
            for i, v in enumerate(s7d):
                ws.cell(row=27, column=3 + i, value=v)

            # 28-day strengths → row 27, columns F-H (6-8)
            for i, v in enumerate(s28d):
                ws.cell(row=27, column=6 + i, value=v)

            total += 1
            log(f"    ✓ {sheets[si]} filled")

        if progress_cb:
            progress_cb((gi + 1) / grade_count * 0.8)

    return total


def apply_generated_grades_from_template(office_wb, log, progress_cb=None):
    """
    Auto mode: read each sheet B12, detect grade/type, generate one row,
    and write directly into that sheet.
    """
    total = 0

    supported_sheets = []
    for sheet_name in office_wb.sheetnames:
        ws = office_wb[sheet_name]
        grade = _grade_from_template_cell(ws["B12"].value)
        if grade:
            supported_sheets.append((sheet_name, grade))

    total_supported = len(supported_sheets)
    log(f"  Supported sheets detected from B12: {total_supported}")

    if total_supported == 0:
        log("  ⚠ No supported grades/types found in template B12 cells")
        return 0

    for i, (sheet_name, grade) in enumerate(supported_sheets):
        ws = office_wb[sheet_name]
        weights, s7d, s28d = next(generate_rows(grade, 1))

        for idx, value in enumerate(weights):
            ws.cell(row=25, column=3 + idx, value=value)
        for idx, value in enumerate(s7d):
            ws.cell(row=27, column=3 + idx, value=value)
        for idx, value in enumerate(s28d):
            ws.cell(row=27, column=6 + idx, value=value)

        total += 1
        log(f"    ✓ {sheet_name} filled ({grade_display_name(grade)})")

        if progress_cb:
            progress_cb((i + 1) / total_supported * 0.8)

    return total


# ── Grade processing (from existing Excel files – legacy) ──────────────────

def apply_grade_files(office_wb, grade_files, log, progress_cb=None):
    """Read existing grade Excel files and populate office template (legacy mode)."""
    total = 0
    file_count = len(grade_files)

    for fi, grade_file in enumerate(grade_files):
        grade_wb = _load_workbook(grade_file)
        grade_ws = grade_wb.active
        grade_name = _extract_grade_from_filename(grade_file)

        log(f"\n  File: {os.path.basename(grade_file)}  (grade: {grade_name})")

        # Find last data row
        row = 2
        while grade_ws.cell(row=row, column=2).value not in (None, ""):
            row += 1
        last_row = row - 1
        log(f"  Data rows: {last_row - 1}")

        sheets = _find_sheets_for_grade(office_wb, grade_name, log)
        log(f"  Matching sheets: {len(sheets)}")

        if not sheets:
            grade_wb.close()
            continue

        si = 0
        for r in range(2, last_row + 1):
            if si >= len(sheets):
                log("  ⚠ More data rows than sheets")
                break
            ws = office_wb[sheets[si]]

            weights = [grade_ws.cell(row=r, column=c).value for c in range(2, 8)]
            strengths = [grade_ws.cell(row=r, column=c).value for c in range(9, 15)]

            for i, v in enumerate(weights):
                ws.cell(row=25, column=3 + i, value=v)
            for i, v in enumerate(strengths):
                ws.cell(row=27, column=3 + i, value=v)

            total += 1
            si += 1

        grade_wb.close()

        if progress_cb:
            progress_cb((fi + 1) / file_count * 0.8)

    return total


# ── Main orchestrator ───────────────────────────────────────────────────────

def process(
    office_file,
    output_folder,
    mode,                    # "generate", "grade_files", "date_only", "generate+date", "grade_files+date"
    log,
    selected_grades=None,    # for generate modes
    num_rows=1000,
    grade_files=None,        # for legacy grade-file modes
    calendar_file=None,
    progress_cb=None,
):
    """
    One-shot processing entry point.

    Returns total count of sheet operations performed.
    """
    log(f"\n{'═' * 60}")
    log(f"  MODE: {mode.upper().replace('_', ' ')}")
    log(f"{'═' * 60}")

    # Prepare output
    base = os.path.splitext(os.path.basename(office_file))[0]
    out_name = f"{base}_Processed.xlsx"
    out_path = os.path.join(output_folder, out_name)
    shutil.copy2(office_file, out_path)
    office_wb = _load_workbook(out_path)

    total = 0

    # Calendar
    calendar_data = None
    if "date" in mode:
        calendar_data = load_calendar_data(calendar_file, log)
        if not calendar_data:
            log("✖ Cannot proceed without valid calendar file")
            office_wb.close()
            return 0

    # Grade generation (AIO)
    if "generate" in mode:
        log("\n── GENERATING & APPLYING GRADE DATA ──")
        if selected_grades:
            total += apply_generated_grades(office_wb, selected_grades, num_rows, log, progress_cb)
        else:
            log("  Auto mode: detecting grade/type from each sheet B12")
            total += apply_generated_grades_from_template(office_wb, log, progress_cb)

    # Grade files (legacy)
    if "grade_files" in mode and grade_files:
        log("\n── APPLYING GRADE FILES ──")
        total += apply_grade_files(office_wb, grade_files, log, progress_cb)

    # Dates
    if calendar_data:
        log("\n── APPLYING DATES ──")
        updated = apply_dates(office_wb, calendar_data, log)
        log(f"  Sheets updated with dates: {updated}")

    # Save
    office_wb.save(out_path)
    office_wb.close()

    log(f"\n{'═' * 60}")
    log(f"  ✓ SAVED → {out_path}")
    log(f"{'═' * 60}")

    if progress_cb:
        progress_cb(1.0)

    return total
